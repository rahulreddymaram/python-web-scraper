from __future__ import annotations

import argparse
import logging
import re
import sys
import time
from typing import Any
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


DEFAULT_BASE_URL = "https://books.toscrape.com/catalogue/page-{}.html"
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}
RATING_MAP = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5,
}
OUTPUT_COLUMNS = [
    "product_name",
    "price_gbp",
    "rating",
    "availability",
    "product_url",
]


def build_session() -> requests.Session:
    session = requests.Session()
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[403, 429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(DEFAULT_HEADERS)
    return session


def fetch_page(url: str, session: requests.Session, timeout: int = 20) -> str | None:
    try:
        response = session.get(url, timeout=timeout)
        response.raise_for_status()
        return response.text
    except requests.RequestException as exc:
        logging.error("Failed to fetch %s: %s", url, exc)
        return None


def parse_data(html: str, page_url: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    products: list[dict[str, Any]] = []

    for product in soup.select("article.product_pod"):
        title_tag = product.select_one("h3 a")
        price_tag = product.select_one("p.price_color")
        availability_tag = product.select_one("p.instock.availability")
        rating_tag = product.select_one("p.star-rating")

        relative_url = title_tag["href"] if title_tag and title_tag.has_attr("href") else ""
        rating_classes = rating_tag.get("class", []) if rating_tag else []
        rating_label = next(
            (name for name in rating_classes if name in RATING_MAP),
            None,
        )

        products.append(
            {
                "product_name": title_tag.get("title", "").strip() if title_tag else "",
                "price_gbp": price_tag.get_text(strip=True) if price_tag else None,
                "rating": RATING_MAP.get(rating_label),
                "availability": availability_tag.get_text(" ", strip=True) if availability_tag else None,
                "product_url": urljoin(page_url, relative_url),
            }
        )

    return products


def clean_data(dataframe: pd.DataFrame) -> pd.DataFrame:
    if dataframe.empty:
        return dataframe.reindex(columns=OUTPUT_COLUMNS)

    cleaned = dataframe.copy()

    for column in ["product_name", "availability", "product_url"]:
        cleaned[column] = (
            cleaned[column]
            .fillna("")
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

    cleaned["price_gbp"] = (
        cleaned["price_gbp"]
        .fillna("")
        .astype(str)
        .str.replace(r"[^0-9.]", "", regex=True)
    )
    cleaned["price_gbp"] = pd.to_numeric(cleaned["price_gbp"], errors="coerce")
    cleaned["rating"] = pd.to_numeric(cleaned["rating"], errors="coerce").astype("Int64")

    cleaned = cleaned.drop_duplicates(subset=["product_url"]).reset_index(drop=True)
    cleaned["availability"] = cleaned["availability"].replace("", "Unknown")
    cleaned = cleaned.reindex(columns=OUTPUT_COLUMNS)
    return cleaned


def save_to_excel(dataframe: pd.DataFrame, output_path: str) -> None:
    dataframe.to_excel(output_path, index=False)

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.title = "Scraped Data"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        adjusted_width = min(max(len(value) for value in values) + 2, 60)
        worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    workbook.save(output_path)


def save_to_csv(dataframe: pd.DataFrame, output_path: str) -> str:
    csv_path = re.sub(r"\.xlsx$", ".csv", output_path, flags=re.IGNORECASE)
    if csv_path == output_path:
        csv_path = f"{output_path}.csv"
    dataframe.to_csv(csv_path, index=False)
    return csv_path


def scrape_products(
    base_url: str,
    pages: int,
    delay: float,
    session: requests.Session,
    timeout: int,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []

    for page_number in range(1, pages + 1):
        page_url = base_url.format(page_number)
        logging.info("Scraping page %s: %s", page_number, page_url)
        html = fetch_page(page_url, session=session, timeout=timeout)

        if not html:
            logging.warning("Skipping page %s because no HTML was returned.", page_number)
            continue

        page_records = parse_data(html, page_url)
        logging.info("Extracted %s records from page %s.", len(page_records), page_number)
        records.extend(page_records)

        if page_number < pages:
            time.sleep(delay)

    dataframe = pd.DataFrame(records)
    return clean_data(dataframe)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape product data from a paginated public website and export it to Excel."
    )
    parser.add_argument(
        "--base-url",
        default=DEFAULT_BASE_URL,
        help="Paginated URL template containing '{}' for the page number.",
    )
    parser.add_argument(
        "--pages",
        type=int,
        default=3,
        help="Number of pages to scrape.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.5,
        help="Delay in seconds between requests.",
    )
    parser.add_argument(
        "--output",
        default="scraped_data.xlsx",
        help="Excel output path.",
    )
    parser.add_argument(
        "--csv",
        action="store_true",
        help="Also export the cleaned data as CSV.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=20,
        help="Request timeout in seconds.",
    )
    args = parser.parse_args()

    if "{}" not in args.base_url:
        parser.error("--base-url must contain '{}' so page numbers can be inserted.")
    if args.pages < 1:
        parser.error("--pages must be at least 1.")
    if args.delay < 0:
        parser.error("--delay cannot be negative.")
    if args.timeout < 1:
        parser.error("--timeout must be at least 1 second.")

    return args


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )


def main() -> int:
    configure_logging()
    args = parse_args()

    session = build_session()

    dataframe = scrape_products(
        base_url=args.base_url,
        pages=args.pages,
        delay=args.delay,
        session=session,
        timeout=args.timeout,
    )

    if dataframe.empty:
        logging.warning("No records were scraped. No output files were created.")
        return 1

    save_to_excel(dataframe, args.output)
    logging.info("Saved Excel output to %s", args.output)

    if args.csv:
        csv_path = save_to_csv(dataframe, args.output)
        logging.info("Saved CSV output to %s", csv_path)

    logging.info("Total cleaned records scraped: %s", len(dataframe))
    return 0
if __name__ == "__main__":
    sys.exit(main())
